import openpyxl;

wb = openpyxl.load_workbook(filename = 'test/sample.xlsx')

wb.create_sheet('sid1')
wb.cell(row=5,column=1,value='Pytest')

wb.save('test/sample.xlsx')